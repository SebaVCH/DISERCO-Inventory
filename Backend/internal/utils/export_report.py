import io
from pathlib import Path
from typing import List, Any

from openpyxl import load_workbook, Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Side, Border
from sqlalchemy.orm import Session
from sqlalchemy import func, distinct, case

from internal.domain.inventory_movement import InventoryMovement
from internal.domain.report import Report
from internal.domain.report_inventory_item import ReportInventoryItem
from internal.domain.inventory_item import InventoryItem
from internal.domain.section import Section

TEMPLATE_PATH = Path(__file__).resolve().parent / "Plantilla_Reporte_Abastecimiento.xlsx"

def export_report_to_excel(report: Report, items_report: List[ReportInventoryItem], db: Session) -> io.BytesIO:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Plantilla no encontrada en {TEMPLATE_PATH}")

    workbook = load_workbook(TEMPLATE_PATH)
    sections = get_sections(items_report)
    side_style = Side(border_style="thin", color="000000")
    cell_border = Border(
        left=side_style,
        right=side_style,
        top=side_style,
        bottom=side_style
    )
    replace_general_summary(report, workbook, db, sections, cell_border)
    generate_total_inventory_detail(items_report, report, workbook, cell_border, db)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer

def generate_total_inventory_detail(items_report: list[ReportInventoryItem], report: Report, workbook: Workbook, cell_border: Border, db: Session):

    sheet_detail = workbook["Inventario Detallado"]
    sheet_detail["C5"] = f"{report.period_end}"

    items_by_id: dict[int, ReportInventoryItem] = {}
    for report_item in items_report:
        inventory_item = report_item.inventory_item
        if inventory_item:
            items_by_id[inventory_item.id] = report_item

    if not items_by_id:
        return

    item_ids = list(items_by_id.keys())

    movements = (
        db.query(InventoryMovement)
        .join(InventoryMovement.inventory_item)
        .outerjoin(InventoryItem.section)
        .filter(InventoryMovement.inventory_item_id.in_(item_ids))
        .filter(InventoryMovement.created_at >= report.period_start)
        .filter(InventoryMovement.created_at <= report.period_end)
        .all()
    )

    movements_after_period = (
        db.query(InventoryMovement)
        .filter(InventoryMovement.inventory_item_id.in_(item_ids))
        .filter(InventoryMovement.created_at > report.period_end)
        .all()
    )

    aggregates: dict[int, dict[str, Any]] = {}
    for mv in movements:
        entry = aggregates.setdefault(mv.inventory_item_id, {
            "entries": 0,
            "exits": 0,
            "observations": set(),
            "responsables": set(),
        })
        if mv.movement_type == "Entrada":
            entry["entries"] += mv.quantity
        elif mv.movement_type == "Salida":
            entry["exits"] += mv.quantity
        if mv.observation:
            entry["observations"].add(mv.observation)
        if mv.user:
            entry["responsables"].add(mv.user.full_name)

    adjustments_after: dict[int, dict[str, Any]] = {}
    for mv in movements_after_period:
        entry = adjustments_after.setdefault(mv.inventory_item_id, {"entries": 0, "exits": 0})
        if mv.movement_type == "Entrada":
            entry["entries"] += mv.quantity
        elif mv.movement_type == "Salida":
            entry["exits"] += mv.quantity

    start_row = 9
    for row_offset, item_id in enumerate(sorted(items_by_id.keys())):
        report_item = items_by_id[item_id]
        item = report_item.inventory_item
        agg = aggregates.get(item_id, {"entries": 0, "exits": 0, "observations": set(), "responsables": set()})
        outside = adjustments_after.get(item_id, {"entries": 0, "exits": 0})

        section = item.section if item else None
        section_name = section.name if section else "Sin Sección"
        base_stock = item.current_stock
        calculated_stock = base_stock - outside["entries"] + outside["exits"]

        row = start_row + row_offset
        sheet_detail.cell(row=row, column=1, value=item.id).border = cell_border
        sheet_detail.cell(row=row, column=2, value=item.name).border = cell_border
        sheet_detail.cell(row=row, column=3, value=section_name).border = cell_border
        sheet_detail.cell(row=row, column=4, value=agg["entries"]).border = cell_border
        sheet_detail.cell(row=row, column=5, value=agg["exits"]).border = cell_border
        sheet_detail.cell(row=row, column=6, value=calculated_stock).border = cell_border
        sheet_detail.cell(row=row, column=7, value="; ".join(sorted(agg["observations"])) if agg["observations"] else None).border = cell_border
        sheet_detail.cell(row=row, column=8, value=", ".join(sorted(agg["responsables"])) if agg["responsables"] else None).border = cell_border

    last_row = start_row + len(items_by_id) - 1
    full_range = f"A8:H{last_row}"
    sheet_detail.auto_filter.ref = full_range

def get_sections(items_report: list[ReportInventoryItem]) -> list[Any]:
    sections = []
    for item in items_report:
        inventory_item = item.inventory_item
        section = inventory_item.section if inventory_item else None
        section_name = section.name if section else None
        if not section_name:
            if "Sin Sección" not in sections:
                sections.append("Sin Sección")
        elif section_name not in sections:
            sections.append(section_name)
    return sections

def replace_general_summary(report: Report, workbook: Workbook, db: Session, sections: List[str], cell_border: Border):
    sheet = workbook["Resumen General"]

    sheet["E4"] = report.id
    sheet["B4"] = f"{report.period_start} - {report.period_end}"
    sheet["B5"] = report.generated_at
    sheet["B6"] = report.user.full_name

    base_query = db.query(ReportInventoryItem).join(ReportInventoryItem.inventory_item).outerjoin(InventoryItem.section).filter(ReportInventoryItem.report_id == report.id)
    report_items_subq = db.query(ReportInventoryItem.inventory_item_id).filter(ReportInventoryItem.report_id == report.id).subquery()

    movements_base = (
        db.query(InventoryMovement).join(InventoryMovement.inventory_item).outerjoin(InventoryItem.section)
        .filter(InventoryMovement.inventory_item_id.in_(report_items_subq))
        .filter(InventoryMovement.created_at >= report.period_start)
        .filter(InventoryMovement.created_at <= report.period_end)
    )

    for row_index, section_name in enumerate(sections, start=1):
        section_query = base_query
        section_movements = movements_base
        if section_name == "Sin Sección":
            section_query = section_query.filter(Section.id.is_(None))
            section_movements = section_movements.filter(Section.id.is_(None))
        else:
            section_query = section_query.filter(Section.name == section_name)
            section_movements = section_movements.filter(Section.name == section_name)

        unique_items = section_query.with_entities(func.count(distinct(InventoryItem.id))).scalar() or 0
        total_movements = section_movements.with_entities(func.count(InventoryMovement.id)).scalar() or 0
        total_entries = section_movements.with_entities(func.coalesce(func.sum(case((InventoryMovement.movement_type == "Entrada", InventoryMovement.quantity), else_=0)),0,)).scalar()or 0
        total_exits = section_movements.with_entities(func.coalesce(func.sum(case((InventoryMovement.movement_type == "Salida", InventoryMovement.quantity), else_=0)),0,)).scalar()or 0

        sheet.cell(row=row_index + 9, column=1, value=section_name).border = cell_border
        sheet.cell(row=row_index + 9, column=2, value=unique_items).border = cell_border
        sheet.cell(row=row_index + 9, column=3, value=total_movements).border = cell_border
        sheet.cell(row=row_index + 9, column=4, value=total_entries).border = cell_border
        sheet.cell(row=row_index + 9, column=5, value=total_exits).border = cell_border